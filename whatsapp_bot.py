import pandas as pd
import random
import os
import time
import threading
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from flask import Flask, request, jsonify
from flask_cors import CORS

class IDecorBot:
    def __init__(self, excel_file="orders.xlsx"):
        self.excel_file = excel_file
        self.user_sessions = {}
        self.ensure_excel_and_columns()
        
        # Policy and text constants
        self.WELCOME_MESSAGE = (
            "Welcome to iDecor 🖼️\n"
            "Custom posters, polaroids & stickers.\n"
            "How can I help you today?\n\n"
            "• 🛒 Place an order\n"
            "• 📦 Check order status\n"
            "• ℹ️ FAQ\n"
            "• 📜 Policies\n"
            "• 🌐 Website link"
        )
        
        self.FAQ_MESSAGE = "Please note: iDecor does not offer replacement. Courier damage is not refundable."

        self.POLICIES_MESSAGE = (
            "📜 iDecor Store Policies\n\n"
            "• All products are customized. Once an order is confirmed, it cannot be modified or cancelled after printing or packaging starts.\n"
            "• Product images are for reference only. Minor color or size variations may occur and are not considered defects.\n"
            "• Refunds are not guaranteed and are issued only after internal verification and approval.\n"
            "• Courier damage after dispatch is not refundable. Packing proof is recorded before dispatch.\n"
            "• Any issue must be reported within 24–48 hours of delivery with a clear unedited unboxing video.\n"
            "• Delays, damage, or loss caused by courier partners are not the responsibility of iDecor.\n"
            "• Refusal to accept delivery, incorrect address, or customer dissatisfaction is non-refundable.\n"
            "• Customized or made-to-order products cannot be cancelled once confirmed.\n"
            "• iDecor does not offer replacement under any circumstances.\n\n"
            "By placing an order, you agree to all store policies."
        )

        self.PAYMENT_QR_MESSAGE = (
            "Please scan the QR code below to complete payment 💳\n"
            "Our team will verify the payment within 7 hours.\n"
            "You will receive a confirmation message after verification."
        )

        self.PAYMENT_PENDING_MESSAGE = (
            "Your payment is under verification ⏳\n"
            "Our team will confirm it within 7 hours."
        )

    def ensure_excel_and_columns(self):
        # Base columns from user's provided schema
        required_columns = [
            'Order ID', 'Customer Name', 'Product Name', 'Product Type', 
            'Size', 'Quantity', 'Order Date', 'address', 
            'Contact no.', 'another contact no.', 'Payment Verified', 'Confirmation Sent'
        ]
        
        if not os.path.exists(self.excel_file):
            df = pd.DataFrame(columns=required_columns)
            df.to_excel(self.excel_file, index=False)
            self.apply_data_validation()
        else:
            # Ensure new columns exist if file already exists
            try:
                df = pd.read_excel(self.excel_file)
                changed = False
                for col in required_columns:
                    if col not in df.columns:
                        df[col] = "" # Add missing column
                        changed = True
                if changed:
                    df.to_excel(self.excel_file, index=False)
                    self.apply_data_validation()
            except Exception as e:
                print(f"Error initializing excel: {e}")

    def apply_data_validation(self):
        """Adds a Yes/No dropdown to the 'Payment Verified' column."""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Create Data Validation for "Yes,No"
            dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            dv.error = 'Your input is not valid'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = 'Please select from the list'
            dv.promptTitle = 'Payment Verification'
            
            # Find the confirmation column index dynamically
            payment_col_letter = "K" # Default fallback
            for cell in ws[1]:
                if cell.value == "Payment Verified":
                    payment_col_letter = cell.column_letter
                    break
            
            ws.add_data_validation(dv)
            
            # Apply to the range
            max_r = ws.max_row
            if max_r < 2: max_r = 2
            dv.add(f"{payment_col_letter}2:{payment_col_letter}{max_r+100}")
            
            wb.save(self.excel_file)
        except Exception as e:
            pass

    def generate_order_id(self):
        return f"ID{random.randint(1000, 9999)}"

    def get_order_status(self, order_id):
        if not os.path.exists(self.excel_file):
            return None
        try:
            df = pd.read_excel(self.excel_file, dtype={'Order ID': str})
            # Ensure comparison as string
            order_row = df[df['Order ID'].astype(str) == str(order_id)]
            
            if not order_row.empty:
                verified = str(order_row.iloc[0].get('Payment Verified', '')).strip().lower()
                if verified == 'yes':
                    return "Confirmed"
                else:
                    return "Payment Pending"
        except Exception:
            pass
        return None

    def save_order(self, order_data):
        try:
            # Reload to get current state
            if os.path.exists(self.excel_file):
                df = pd.read_excel(self.excel_file)
            else:
                df = pd.DataFrame()
            
            # Map bot data keys to Excel columns
            row_data = {
                'Order ID': order_data['order_id'],
                'Customer Name': order_data['name'],
                'Product Name': order_data['product_name'],
                'Product Type': order_data['type'],
                'Size': order_data['size'],
                'Quantity': order_data['qty'],
                'Order Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'address': '', # Not collected per prompt
                'Contact no.': order_data['phone'],
                'another contact no.': '',
                'Payment Verified': 'No',
                'Confirmation Sent': 'No'
            }
            
            new_row = pd.DataFrame([row_data])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(self.excel_file, index=False)
            
            # Apply dropdowns after saving
            self.apply_data_validation()
            return True
        except Exception as e:
            print(f"Error saving order: {e}")
            return False

    def check_for_notifications(self):
        """Checks for confirmed payments that haven't been notified yet."""
        if not os.path.exists(self.excel_file):
            return []
        
        notifications = []
        try:
            # Re-read file to get latest updates
            df = pd.read_excel(self.excel_file)
            
            # Check if columns exist
            required_cols = ['Payment Verified', 'Confirmation Sent', 'Order ID']
            if any(col not in df.columns for col in required_cols):
                return []

            # Helper to normalize 'Yes'/'No'
            def is_yes(val):
                return str(val).strip().lower() == 'yes'

            # Find rows: Payment=Yes AND Confirmation!=Yes
            mask = df.apply(lambda x: is_yes(x.get('Payment Verified')) and not is_yes(x.get('Confirmation Sent')), axis=1)
            
            rows_to_notify = df[mask]
            
            for index, row in rows_to_notify.iterrows():
                phone = str(row.get('Contact no.', 'Unknown'))
                order_id = row['Order ID']
                notifications.append((phone, f"Your order #{order_id} is Confirmed! ✅"))
                
                # Mark as notified
                df.at[index, 'Confirmation Sent'] = 'Yes'
            
            if not rows_to_notify.empty:
                df.to_excel(self.excel_file, index=False)
                self.apply_data_validation()
                
        except PermissionError:
            return ["ERROR: Could not read Excel file. Please close it if it is open!"]
        except Exception as e:
            pass
            
        return notifications

    def handle_message(self, user_phone, message):
        message = message.strip()
        
        if user_phone not in self.user_sessions:
            self.user_sessions[user_phone] = {"state": "IDLE", "data": {}}
        
        session = self.user_sessions[user_phone]
        state = session["state"]
        
        # Global resets
        if message.lower() in ["hi", "hello", "hey"]:
            session["state"] = "IDLE"
            session["data"] = {}
            return self.WELCOME_MESSAGE

        # IDLE
        if state == "IDLE":
            if "place an order" in message.lower():
                session["state"] = "ASK_NAME"
                return "Please enter your name:"
            
            elif "check order status" in message.lower():
                session["state"] = "CHECK_STATUS"
                return "Please enter your Order ID:"
            
            elif "faq" in message.lower():
                return self.FAQ_MESSAGE
            
            elif "policies" in message.lower() or any(x in message.lower() for x in ["refund", "cancellation", "store rules", "replacement"]):
                return self.POLICIES_MESSAGE
            
            elif "website link" in message.lower():
                return "🌐 Visit us at: https://idecor.com"
            
            elif "cancel" in message.lower() and "order" in message.lower():
                 return "Orders can only be cancelled before printing or packaging begins. Once processing starts, cancellation is not possible."

            elif "confirmed" in message.lower() and ("payment" in message.lower() or "order" in message.lower()):
                return self.PAYMENT_PENDING_MESSAGE
            
            else:
                return "I didn't understand that. Type 'hi' to see the options."

        # ORDER COLLECTION FLOW
        elif state == "ASK_NAME":
            session["data"]["name"] = message
            session["state"] = "ASK_PRODUCT"
            return "Please type the product name exactly as you want it:"

        elif state == "ASK_PRODUCT":
            session["data"]["product_name"] = message
            session["state"] = "ASK_TYPE"
            return "Select a product type:\n• Custom poster\n• Custom polaroid\n• Custom sticker"

        elif state == "ASK_TYPE":
            msg_lower = message.lower()
            if "poster" in msg_lower:
                session["data"]["type"] = "Custom poster"
                session["state"] = "ASK_SIZE_POSTER"
                return "Which size would you like? A4 / A3 / 12x18 / 13x19"
            elif "polaroid" in msg_lower:
                session["data"]["type"] = "Custom polaroid"
                session["data"]["size"] = "NA"
                session["state"] = "ASK_QTY"
                return "Please enter the quantity:"
            elif "sticker" in msg_lower:
                session["data"]["type"] = "Custom sticker"
                session["state"] = "ASK_SIZE_STICKER"
                return "Please enter the sticker size:"
            else:
                return "Please select a valid type (Custom poster, Custom polaroid, Custom sticker)."

        elif state == "ASK_SIZE_POSTER":
            valid_sizes = ["a4", "a3", "12x18", "13x19"]
            if message.lower() not in valid_sizes:
                return "Please choose a valid size: A4 / A3 / 12x18 / 13x19"
            session["data"]["size"] = message.upper()
            session["state"] = "ASK_QTY"
            return "Please enter the quantity:"

        elif state == "ASK_SIZE_STICKER":
            session["data"]["size"] = message
            session["state"] = "ASK_QTY"
            return "Please enter the quantity:"

        elif state == "ASK_QTY":
            if not message.isdigit():
                return "Please enter a valid number for quantity."
            
            session["data"]["qty"] = int(message)
            session["data"]["phone"] = user_phone
            
            # Generate Order ID & Save
            order_id = self.generate_order_id()
            session["data"]["order_id"] = order_id
            
            if self.save_order(session["data"]):
                session["state"] = "IDLE"
                receipt = (
                    f"Your order has been booked ✅\n"
                    f"Order ID: #{order_id}\n"
                    f"Product: {session['data']['product_name']}\n"
                    f"Type: {session['data']['type']}\n"
                    f"Size: {session['data']['size']}\n"
                    f"Quantity: {session['data']['qty']}\n"
                    f"Status: Payment Pending\n\n"
                    f"{self.PAYMENT_QR_MESSAGE}"
                )
                
                # SImulate sending image
                return receipt + "\n\n[BOT: Sending QR Code Image...]"
            else:
                return "Error saving order. Please try again."

        # STATUS CHECK
        elif state == "CHECK_STATUS":
            order_id_input = message.replace("#", "").strip()
            if "cancel" in message.lower():
                session["state"] = "IDLE"
                return "Orders can only be cancelled before printing or packaging begins. Once processing starts, cancellation is not possible."

            status = self.get_order_status(order_id_input)
            session["state"] = "IDLE"
            
            if status:
                return status
            else:
                return "Order ID not found."

        return "I didn't understand that. Type 'hi' to start over."

def monitor_notifications(bot):
    """Background thread to check for notifications."""
    print("\n[SYSTEM]: Background monitoring initialized.")
    while True:
        results = bot.check_for_notifications()
        
        # --- VERBOSE DEBUGGING FOR USER ---
        # Show what the bot sees in the file every few loops
        try:
             if os.path.exists(bot.excel_file):
                # We check silently just to help user debug
                # Note: This adds IO overhead but is necessary for troubleshooting
                df_debug = pd.read_excel(bot.excel_file)
                # Find pending orders
                pending = df_debug[df_debug['Confirmation Sent'].astype(str).str.lower().str.strip() != 'yes']
                # Check their payment status
                if not pending.empty:
                    # just verify the last one
                    last_row = pending.iloc[-1]
                    ver_status = str(last_row['Payment Verified'])
                    order_id = str(last_row['Order ID'])
                    # If user is staring at screen waiting...
                    if ver_status.lower().strip() == 'yes':
                        # This should have been caught by 'results' logic above, 
                        # but if we are here, something was missed or race condition.
                        pass
                    else:
                        print(f"[Checking Excel]: Order {order_id} has Payment Verified='{ver_status}'. Waiting for 'Yes'...", end="\r")
        except:
            pass
        # ----------------------------------

        for res in results:
            if isinstance(res, str) and "ERROR" in res:
                # Only print error once per occurrence
                print(f"\n[SYSTEM]: {res}\nUser: ", end="")
            elif isinstance(res, tuple):
                phone, msg = res
                print(f"\n\n[WHATSAPP NOTIFICATION to {phone}]: {msg}\nUser: ", end="")
        
        time.sleep(3) # Check every 3 seconds

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

bot = IDecorBot()

# Start background thread for notifications (logs to server console)
# We check if it is the main thread/reloader to avoid duplicate threads in some envs, 
# but for simple deployment this is okay.
t = threading.Thread(target=monitor_notifications, args=(bot,), daemon=True)
t.start()

@app.route('/')
def home():
    return app.send_static_file('index.html')

@app.route('/chat', methods=['POST'])
def chat():
    data = request.json
    message = data.get('message', '')
    user_id = data.get('user_id', 'web_guest') # Use provided user_id or default
    
    response = bot.handle_message(user_id, message)
    return jsonify({"response": response})

if __name__ == "__main__":
    print("Starting Flask connection for iDecor Chat...")
    app.run(port=5000, debug=True)
